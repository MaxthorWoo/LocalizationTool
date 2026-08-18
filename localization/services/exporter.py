"""导出服务：将翻译校对结果导出为 xlsx 表格文件。

导出列设计：序号、源文案、各目标语言译文（每语言一列）、状态、命中术语。
使用 openpyxl 精细控制样式（专业字体、表头、列宽、状态徽标色）。
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .. import config
from ..models import STATUS_LABELS
from .parser import _clean_cell

# 状态 -> 徽标填充色（浅底深字）
STATUS_COLORS = {
    "pending": "E2E8F0",     # 灰
    "translated": "DBEAFE",  # 蓝
    "proofread": "D1FAE5",   # 绿
    "review": "FEF3C7",      # 橙
}

# 语言代码 -> 显示名（导出表头使用）
LANG_NAMES = {
    "zh-CN": "简体中文",
    "zh-TW": "繁體中文",
    "en": "English",
    "ja": "日本語",
    "ko": "한국어",
    "th": "ไทย",
    "id": "Bahasa Indonesia",
    "vi": "Tiếng Việt",
    "fr": "Français",
    "de": "Deutsch",
    "es": "Español",
    "pt": "Português",
    "ru": "Русский",
    "ar": "العربية",
}

HEADER_FILL = PatternFill("solid", start_color="2563EB")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=11)
WRAP = Alignment(vertical="top", wrap_text=True)


def export_project(
    project,
    entries: list,
    export_dir: str | None = None,
    target_langs: list[str] | None = None,
) -> str:
    """将工程的条目列表导出为 xlsx，返回导出文件路径。

    target_langs 优先取调用方传入的语言方案列表；为空时回退工程的 target_langs。
    """
    export_dir = export_dir or str(config.EXPORT_DIR)
    os.makedirs(export_dir, exist_ok=True)

    target_langs = target_langs or project.get_target_lang_list()

    wb = Workbook()
    ws = wb.active
    ws.title = "翻译结果"

    # 表头
    headers = ["序号", "源文案", "键"]
    headers.extend(LANG_NAMES.get(lang, lang) for lang in target_langs)
    headers.extend(["状态", "命中术语"])
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 数据行
    for idx, entry in enumerate(entries, start=1):
        row = [idx, _clean_cell(entry.source_text), _clean_cell(entry.key_text)]
        trans = entry.get_translations()
        for lang in target_langs:
            row.append(trans.get(lang, ""))
        status_label = STATUS_LABELS.get(entry.status, entry.status)
        row.append(status_label)
        row.append(", ".join(entry.get_term_hits()) if entry.get_term_hits() else "")
        ws.append(row)
        # 样式
        excel_row = idx + 1
        status_cell = ws.cell(row=excel_row, column=len(headers) - 1)
        color = STATUS_COLORS.get(entry.status)
        if color:
            status_cell.fill = PatternFill("solid", start_color=color)
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=excel_row, column=col_idx).font = BODY_FONT
            ws.cell(row=excel_row, column=col_idx).alignment = WRAP

    # 列宽
    col_widths = {1: 6, 2: 40, 3: 12}
    for i, lang in enumerate(target_langs, start=4):
        col_widths[i] = 36
    col_widths[len(headers) - 1] = 12
    col_widths[len(headers)] = 20
    for col_idx, width in col_widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.freeze_panes = "A2"

    filename = f"{_safe_name(project.name)}_translation.xlsx"
    out_path = os.path.join(export_dir, filename)
    wb.save(out_path)
    return out_path


def _safe_name(name: str) -> str:
    """生成安全文件名（去非法字符）。"""
    return "".join(c for c in name if c.isalnum() or c in "._- ").strip() or "project"
